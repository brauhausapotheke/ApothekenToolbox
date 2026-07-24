from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


GRUEN = PatternFill(fill_type="solid", fgColor="C6EFCE")
GELB = PatternFill(fill_type="solid", fgColor="FFF2CC")
ROT = PatternFill(fill_type="solid", fgColor="F4CCCC")


def erstelle_trefferanalyse(artikel_liste):

    wb = Workbook()
    ws = wb.active

    ws.title = "Trefferanalyse"

    kopf = [
        "Ornamentum",
        "Normalisiert",
        "Tabelle",
        "Treffer",
        "Trefferart",
        "Score",
        "Status",
        "AVS",
        "PZN",
        "AVP",
        "AEP",
        "Kandidat 1",
        "Score 1",
        "Kandidat 2",
        "Score 2",
        "Kandidat 3",
        "Score 3"
    ]

    for spalte, text in enumerate(kopf, start=1):
        zelle = ws.cell(row=1, column=spalte)
        zelle.value = text
        zelle.font = Font(bold=True)

    exakte = 0
    fuzzy = 0
    offen = 0
    kandidaten = 0
    ohne_kandidat = 0

    zeile = 2

    for artikel in artikel_liste:

        if artikel.treffer:

            if artikel.trefferart == "EXAKT":
                exakte += 1
                farbe = GRUEN
            else:
                fuzzy += 1
                farbe = GELB

            status = "Treffer"

        else:

            offen += 1
            farbe = ROT

            if artikel.kandidaten:
                kandidaten += 1
                status = "Kandidat vorhanden"
            else:
                ohne_kandidat += 1
                status = "Kein Kandidat"

        ws.cell(row=zeile, column=1).value = artikel.name
        ws.cell(row=zeile, column=2).value = artikel.normalisiert
        ws.cell(row=zeile, column=3).value = artikel.tabelle
        ws.cell(row=zeile, column=4).value = "JA" if artikel.treffer else "NEIN"
        ws.cell(row=zeile, column=5).value = artikel.trefferart
        ws.cell(row=zeile, column=6).value = round(artikel.fuzzy_score * 100, 1) if artikel.fuzzy_score else ""
        ws.cell(row=zeile, column=7).value = status

        if artikel.treffer:

            ws.cell(row=zeile, column=8).value = artikel.avs_name
            ws.cell(row=zeile, column=9).value = artikel.pzn
            ws.cell(row=zeile, column=10).value = artikel.avp
            ws.cell(row=zeile, column=11).value = artikel.aep

        else:

            spalte = 12

            for score, kandidat in artikel.kandidaten:

                ws.cell(row=zeile, column=spalte).value = kandidat.name
                ws.cell(row=zeile, column=spalte + 1).value = round(score * 100, 1)

                spalte += 2

        for sp in range(1, 18):
            ws.cell(row=zeile, column=sp).fill = farbe

        zeile += 1

    ws.auto_filter.ref = ws.dimensions

    breiten = {
        "A": 50,
        "B": 45,
        "C": 20,
        "D": 10,
        "E": 12,
        "F": 10,
        "G": 22,
        "H": 50,
        "I": 15,
        "J": 12,
        "K": 12,
        "L": 50,
        "M": 10,
        "N": 50,
        "O": 10,
        "P": 50,
        "Q": 10,
    }

    for spalte, breite in breiten.items():
        ws.column_dimensions[spalte].width = breite

    # Zusammenfassung oberhalb der Tabelle
    ws.insert_rows(1, amount=7)

    ws["A1"] = "Trefferanalyse"
    ws["A1"].font = Font(bold=True, size=16)

    ws["A3"] = "Exakte Treffer"
    ws["B3"] = exakte

    ws["A4"] = "Fuzzy Treffer"
    ws["B4"] = fuzzy

    ws["A5"] = "Offen"
    ws["B5"] = offen

    ws["A6"] = "Davon Kandidat vorhanden"
    ws["B6"] = kandidaten

    ws["A7"] = "Davon kein Kandidat"
    ws["B7"] = ohne_kandidat

    dateiname = "Trefferanalyse.xlsx"

    try:
        wb.save(dateiname)
        print()
        print(f"Report gespeichert: {dateiname}")

    except PermissionError:
        print()
        print("❌ Trefferanalyse.xlsx ist noch geöffnet.")
        print("Bitte Excel schließen und erneut starten.")